"""Camada de serviços da app 'consulta'.

Responsabilidades principais:
- Sanitização e formatação de CNPJ.
- Integração com CNPJÁ PRO via `clients.cnpja.CNPJAClient` (com retry/backoff).
- Processamento de entradas CSV/XLSX e agregação dos resultados.
- Exportação em formatos CSV/XLSX.
"""

import re
import csv
import io
import time
import requests
import openpyxl
import xlsxwriter
from django.conf import settings
from clients.cnpja import CNPJAClient, CNPJAClientError
from django.core.cache import cache

DELAY_SECONDS = 0.5  # Delay fixo entre consultas para respeitar limites de API e UX (0.5s)


# Rate limit simples (janela fixa) para no máx. 60 chamadas/minuto à API externa.
# Usa o cache configurado (idealmente Redis em produção) para coordenar entre processos.
RATE_LIMIT_PER_MINUTE = 60
RATE_LIMIT_WINDOW = 60  # segundos

def _rate_limit_acquire(key: str = 'cnpja_api', limit: int = RATE_LIMIT_PER_MINUTE, window_seconds: int = RATE_LIMIT_WINDOW):
    """Bloqueia a chamada até que haja "slot" disponível dentro do limite.

    Implementação com janela fixa por minuto. Em Redis, `incr` é atômico.
    Em LocMemCache, coordena por processo; suficiente em dev.
    """
    now = time.time()
    window = int(now // window_seconds)
    cache_key = f"rl:{key}:{window}"
    # Inicializa contador da janela com TTL restante do minuto
    ttl = window_seconds - int(now % window_seconds) + 1
    if cache.get(cache_key) is None:
        cache.set(cache_key, 0, ttl)
    # Verifica se já atingiu o teto
    try:
        current = cache.get(cache_key) or 0
        if current >= limit:
            # Espera até próxima janela
            wait = window_seconds - (now % window_seconds) + 0.01
            print(f"[RATE LIMIT] Atingido {limit}/min. Aguardando {wait:.2f}s para liberar próximo slot...")
            time.sleep(wait)
            return _rate_limit_acquire(key, limit, window_seconds)
        # Reserva um slot
        try:
            cache.incr(cache_key)
        except Exception:
            cache.set(cache_key, current + 1, ttl)
    except Exception:
        # Em caso de falha no cache, não bloquear a execução (best effort)
        pass


def processar_cnpjs_manualmente(cnpjs: str, on_retry=None):
    """Processa uma string de CNPJs separados por vírgula sequencialmente.

    Retorna (lista_cnpjs_limpos, lista_resultados). Aplica DELAY_SECONDS entre chamadas.
    """
    cnpj_list = [clean_cnpj(c) for c in cnpjs.split(',') if clean_cnpj(c)]
    resultados = []
    for cnpj in cnpj_list:
        resultado = consultar_cnpj_api(cnpj, on_retry=on_retry)
        resultados.append(resultado)
        print(f"[DELAY] Aguardando {DELAY_SECONDS}s para próxima requisição...")
        time.sleep(DELAY_SECONDS)
    return cnpj_list, resultados


def clean_cnpj(cnpj):
    """Normaliza CNPJ:

    - Remove todos os caracteres não numéricos;
    - Caso tenha 12 ou 13 dígitos (zeros à esquerda perdidos), preenche com zeros à esquerda até 14.
    """
    digits = re.sub(r'\D', '', str(cnpj or ''))
    if 12 <= len(digits) < 14:
        digits = digits.zfill(14)
    return digits


def format_cnpj(cnpj):
    """Formata um CNPJ 14 dígitos para 00.000.000/0000-00."""
    cnpj = re.sub(r'\D', '', cnpj)
    if len(cnpj) == 14:
        return f"{cnpj[:2]}.{cnpj[2:5]}.{cnpj[5:8]}/{cnpj[8:12]}-{cnpj[12:]}"
    return cnpj


def consultar_cnpj_api(cnpj, retry_count=3, retry_wait=20, on_retry=None):
    """Consulta a API PRO do CNPJÁ com retry/backoff e extração resiliente de campos.

    - retry_count: tentativas para erros transitórios (429/timeout/connerror).
    - retry_wait: segundos de espera entre tentativas (backoff constante).
    - on_retry: callback opcional (attempt:int, wait:int) para feedback de UI.
    """
    client = CNPJAClient()
    clean = clean_cnpj(cnpj)
    last_error = None
    prefer_cache_first = getattr(settings, 'CNPJA_FORCE_CACHE_FIRST', True)
    # Monta a sequência de estratégias: tenta CACHE puro antes de consultar online
    base_strategy = getattr(settings, 'CNPJA_STRATEGY', 'CACHE_IF_FRESH')
    max_age = getattr(settings, 'CNPJA_MAX_AGE_DAYS', 40)
    max_stale = getattr(settings, 'CNPJA_MAX_STALE_DAYS', 30)
    strategies = []
    if prefer_cache_first:
        strategies.append(('CACHE', None, None))  # não consome créditos
    strategies.append((base_strategy, max_age, max_stale))

    for attempt in range(retry_count):
        # Dentro de cada tentativa, percorre a sequência de estratégias
        for strat, s_max_age, s_max_stale in strategies:
            try:
                # Aplica rate limit apenas quando a estratégia não é puramente de CACHE
                if strat != 'CACHE':
                    _rate_limit_acquire('cnpja_api')
                start_time = time.time()
                data = client.get_office(
                    clean,
                    timeout=30,
                    strategy=strat,
                    max_age_days=s_max_age,
                    max_stale_days=s_max_stale,
                )
                elapsed = time.time() - start_time
                nome = (
                    (data.get('company') or {}).get('name')
                    or data.get('name')
                    or '-'
                )
                email = 'Sem e-mail'
                emails = data.get('emails')
                if isinstance(emails, list) and emails:
                    first = emails[0]
                    if isinstance(first, dict):
                        email = first.get('address') or first.get('email') or 'Sem e-mail'
                    elif isinstance(first, str):
                        email = first
                stale_flag = data.get('stale')
                via = strat + (' (stale)' if stale_flag else '')
                print(f"[CONSULTA] CNPJ {format_cnpj(clean)} | via={via} | resposta={elapsed:.2f}s")
                return {
                    'cnpj': format_cnpj(clean),
                    'nome': nome,
                    'email': email,
                    'detalhes': data
                }
            except CNPJAClientError as e:
                msg = str(e)
                last_error = msg
                # Se estratégia CACHE não encontrou dados (404), tenta próxima sem contar como retry
                if strat == 'CACHE' and '404' in msg:
                    print(f"[CACHE MISS] CNPJ {format_cnpj(clean)}: sem dados em cache. Tentando estratégia {base_strategy}...")
                    continue
                print(f"[ERRO API PRO] via={strat} CNPJ {format_cnpj(clean)}: {msg}")
                if '429' in msg:
                    if on_retry:
                        on_retry(attempt + 1, retry_wait)
                    # passa para próxima tentativa (retry)
                    break
                # Para outros erros, encerra
                return {
                    'cnpj': format_cnpj(clean),
                    'nome': '-',
                    'email': f'Erro API PRO: {msg[:200]}',
                    'detalhes': None
                }
            except (requests.exceptions.Timeout, requests.exceptions.ConnectionError):
                last_error = 'Timeout/ConnectionError'
                print(f"[TIMEOUT/CONNECTION ERROR] via={strat} CNPJ {format_cnpj(clean)}: Tentativa {attempt+1}")
                # passa para próxima tentativa (retry)
                break
            except Exception as e:
                last_error = str(e)
                print(f"[ERRO INESPERADO] CNPJ {format_cnpj(clean)}: {str(e)}")
                return {
                    'cnpj': format_cnpj(clean),
                    'nome': '-',
                    'email': f'Erro inesperado: {str(e)[:200]}',
                    'detalhes': None
                }
        # se chegou aqui, irá para a próxima tentativa por causa de 429/timeout
        if on_retry:
            on_retry(attempt + 1, retry_wait)
    return {
        'cnpj': format_cnpj(clean),
        'nome': '-',
        'email': f'Limite de tentativas excedido: {last_error}',
        'detalhes': None
    }


def _extract_first_cnpj_from_text(text: str):
    """Extrai o primeiro CNPJ (mascarado ou 12–14 dígitos) de um texto; retorna 14 dígitos (zerado à esquerda se necessário) ou None."""
    if not text:
        return None
    # 00.000.000/0000-00
    m = re.search(r"\b\d{2}\.\d{3}\.\d{3}/\d{4}-\d{2}\b", text)
    if m:
        return clean_cnpj(m.group(0))
    # 12 a 14 dígitos (para capturar CNPJs que perderam zeros à esquerda)
    m = re.search(r"(?<!\d)(\d{12,14})(?!\d)", text)
    if m:
        return clean_cnpj(m.group(1))
    return None


def _extract_first_processo_from_text(text: str):
    """Extrai o primeiro Processo: 'xxx.xxx/xxxx' ou 10 dígitos; retorna string ou None."""
    if not text:
        return None
    m = re.search(r"\b\d{3}\.\d{3}/\d{4}\b", text)
    if m:
        return m.group(0)
    m = re.search(r"(?<!\d)(\d{10})(?!\d)", text)
    if m:
        return m.group(1)
    return None


def _safe_decode_bytes(data: bytes) -> str:
    """Decodifica bytes como UTF-8 com fallback para latin-1."""
    try:
        return data.decode('utf-8')
    except UnicodeDecodeError:
        return data.decode('latin-1', errors='ignore')


def format_processo(proc: str | None) -> str | None:
    """Padroniza o número do processo no formato xxx.xxx/xxxx.

    Regras:
    - Se já estiver no padrão xxx.xxx/xxxx, retorna como está (strip antes).
    - Se tiver 10 dígitos, reformatar para xxx.xxx/xxxx.
    - Caso contrário, retorna o valor original com trim (ou None se vazio).
    """
    if proc is None:
        return None
    proc = str(proc).strip()
    if not proc:
        return None
    # Já está no padrão
    if re.fullmatch(r"\d{3}\.\d{3}/\d{4}", proc):
        return proc
    # Somente dígitos (ex.: 8708002017)
    digits = re.sub(r"\D", "", proc)
    if len(digits) == 10:
        return f"{digits[:3]}.{digits[3:6]}/{digits[6:]}"
    return proc


def processar_csv(file, logger=None, on_retry=None):
    """Lê um CSV, detecta colunas e consulta a API por linha.

    - Detecta por nomes de cabeçalho comuns (inclui NRCPFCNPJ/DSProcesso);
    - Fallback: varre a linha por regex para CNPJ e Processo.
    """
    decoded = _safe_decode_bytes(file.read())
    reader = csv.DictReader(io.StringIO(decoded))
    cnpj_keys = ['cnpj', 'CNPJ', 'CNPJ/CPF', 'cnpj_cpf', 'nrcpfcnpj', 'NRCPFCNPJ', 'cpf/cnpj', 'CNPJCPF']
    proc_keys = ['processo', 'Processo', 'número do processo', 'numero do processo', 'dsprocesso', 'DSProcesso']
    resultados = []
    for row in reader:
        cnpj_val = None
        proc_val = None

        # 1) Detectar por cabeçalho
        for key in row:
            if not cnpj_val and any(k.lower() in key.lower() for k in cnpj_keys):
                original = row[key]
                cnpj_val = clean_cnpj(original or '').strip()
                if logger and original:
                    logger.info(f"Linha CSV: '{original}' -> CNPJ limpo: '{cnpj_val}'")
            if not proc_val and any(k.lower() in key.lower() for k in proc_keys):
                proc_val = (row[key] or '').strip()

        # 2) Fallback por regex
        if not cnpj_val:
            joined = ' '.join([str(v) for v in row.values() if v is not None])
            cnpj_val = _extract_first_cnpj_from_text(joined)
        if not proc_val:
            joined = ' '.join([str(v) for v in row.values() if v is not None])
            proc_val = _extract_first_processo_from_text(joined)
        # Padroniza processo
        proc_val = format_processo(proc_val)

        if cnpj_val:
            try:
                if logger:
                    logger.info(f'Consultando CNPJ (CSV): {cnpj_val}')
                resultado = consultar_cnpj_api(cnpj_val, on_retry=on_retry)
                resultado['processo'] = proc_val
                resultados.append(resultado)
            except Exception as e:
                if logger:
                    logger.error(f'Erro ao consultar CNPJ {cnpj_val}: {e}')
                resultados.append({'processo': proc_val, 'cnpj': format_cnpj(cnpj_val), 'nome': '-', 'email': f'Erro: {str(e)}', 'detalhes': None})
            time.sleep(DELAY_SECONDS)
    return resultados


def processar_xlsx(file, logger=None, on_retry=None):
    """Lê um XLSX (primeira planilha), detecta colunas e consulta API por linha.

    Idêntico ao CSV: tenta cabeçalhos, com fallback por regex linha a linha.
    """
    wb = openpyxl.load_workbook(file, read_only=True)
    ws = wb.active
    headers = [str(cell.value).strip() if cell.value else '' for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    cnpj_keys = ['cnpj', 'CNPJ', 'CNPJ/CPF', 'cnpj_cpf', 'nrcpfcnpj', 'NRCPFCNPJ', 'cpf/cnpj', 'CNPJCPF']
    proc_keys = ['processo', 'Processo', 'número do processo', 'numero do processo', 'dsprocesso', 'DSProcesso']
    cnpj_idx = None
    proc_idx = None
    for idx, h in enumerate(headers):
        if cnpj_idx is None and any(k.lower() in h.lower() for k in cnpj_keys):
            cnpj_idx = idx
        if proc_idx is None and any(k.lower() in h.lower() for k in proc_keys):
            proc_idx = idx

    resultados = []
    for row in ws.iter_rows(min_row=2):
        # 1) Por cabeçalho
        cnpj_val = None
        if cnpj_idx is not None and row[cnpj_idx].value:
            cnpj_val = clean_cnpj(str(row[cnpj_idx].value))
        proc_val = (str(row[proc_idx].value).strip() if (proc_idx is not None and row[proc_idx].value) else None)

        # 2) Fallback por regex
        if not cnpj_val:
            joined = ' '.join([str(c.value) for c in row if c.value is not None])
            cnpj_val = _extract_first_cnpj_from_text(joined)
        if not proc_val:
            joined = ' '.join([str(c.value) for c in row if c.value is not None])
            proc_val = _extract_first_processo_from_text(joined)
        # Padroniza processo
        proc_val = format_processo(proc_val)

        if cnpj_val:
            try:
                resultado = consultar_cnpj_api(cnpj_val, on_retry=on_retry)
                resultado['processo'] = proc_val
                resultados.append(resultado)
            except Exception as e:
                resultados.append({'processo': proc_val, 'cnpj': format_cnpj(cnpj_val), 'nome': '-', 'email': f'Erro: {str(e)}', 'detalhes': None})
            time.sleep(DELAY_SECONDS)
    return resultados


def exportar_csv(resultados, include_data=False):
    """Gera CSV em memória a partir da lista de resultados.

    Quando include_data=True, inclui a coluna Data (para histórico).
    """
    output = io.StringIO()
    writer = csv.writer(output)
    if include_data:
        writer.writerow(['Data', 'Processo', 'CNPJ', 'Nome', 'E-mail'])
        for r in resultados:
            writer.writerow([
                r.get('data', ''),
                r.get('processo', ''),
                r.get('cnpj', ''),
                r.get('nome', ''),
                r.get('email', '')
            ])
    else:
        writer.writerow(['Processo', 'CNPJ', 'Nome', 'E-mail'])
        for r in resultados:
            email = r.get('email', '')
            if email in ('', '-', None):
                email = 'Sem e-mail'
            writer.writerow([r.get('processo', ''), r.get('cnpj', ''), r.get('nome', ''), email])
    return output.getvalue()


def exportar_xlsx(resultados, include_data=False):
    """Gera XLSX em memória a partir da lista de resultados.

    Quando include_data=True, inclui a coluna Data (para histórico).
    """
    output = io.BytesIO()
    wb = xlsxwriter.Workbook(output, {'in_memory': True})
    ws = wb.add_worksheet('Export')
    if include_data:
        ws.write_row(0, 0, ['Data', 'Processo', 'CNPJ', 'Nome', 'E-mail'])
        for idx, r in enumerate(resultados, 1):
            ws.write_row(idx, 0, [r.get('data', ''), r.get('processo', ''), r.get('cnpj', ''), r.get('nome', ''), r.get('email', '')])
    else:
        ws.write_row(0, 0, ['Processo', 'CNPJ', 'Nome', 'E-mail'])
        for idx, r in enumerate(resultados, 1):
            email = r.get('email', '')
            if email in ('', '-', None):
                email = 'Sem e-mail'
            ws.write_row(idx, 0, [r.get('processo', ''), r.get('cnpj', ''), r.get('nome', ''), email])
    wb.close()
    output.seek(0)
    return output.read()
