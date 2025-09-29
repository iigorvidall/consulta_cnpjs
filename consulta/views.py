"""Views da aplicação 'consulta'.

Este módulo contém:
- Views HTML (home) com proteção por login.
- Endpoints de exportação de resultados/histórico (CSV/XLSX) baseados em sessão ou banco.
- Endpoints auxiliares (créditos, detalhes por CNPJ) com cache e fallback.
- Um fluxo de processamento em lote baseado em sessão (jobs_* via polling).
- Autenticação (login/logout/signup) com mitigação de brute force via cache.
"""

from django.http import JsonResponse
from django.shortcuts import render
from .models import ConsultaHistorico
import logging
from django.http import HttpResponse
from .services import clean_cnpj, format_cnpj, consultar_cnpj_api, processar_csv, processar_xlsx, exportar_csv, exportar_xlsx, processar_cnpjs_manualmente
from clients.cnpja import CNPJAClient, CNPJAClientError
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from .serializers import CNPJQuerySerializer
from django.views.decorators.http import require_http_methods
from django.views.decorators.csrf import csrf_exempt
from django.utils.decorators import method_decorator
import json
import re
from django.core.cache import cache
from django.views.decorators.http import require_GET
from django.contrib import messages
from django.contrib.auth import authenticate, login as auth_login, logout as auth_logout, get_user_model
from django.contrib.auth.decorators import login_required
from django.shortcuts import redirect
from django.utils import timezone
from .forms import ConsultaForm  # existing
from .forms import LoginForm
from rest_framework.permissions import IsAuthenticated
from django.conf import settings
import os

@login_required(login_url='login')
def status_retry(request):
	"""Retorna o status textual do último retry registrado na sessão.

	Usado pelo frontend para exibir mensagens durante backoff (por exemplo, HTTP 429).
	"""
	status = request.session.get('status_retry', '')
	return JsonResponse({'status': status})

@login_required(login_url='login')
def home(request):
	"""Renderiza a página principal de consulta e trata submissões POST.

	POST aceita:
	- cnpjs: lista separada por vírgulas (processamento manual);
	- csv_file: arquivo CSV ou XLSX (processamento por upload).

	Persiste um snapshot do resultado (ou erro) em `ConsultaHistorico` e
	os últimos resultados na sessão para exportações.
	"""
	error_msg = None
	resultados = []
	historico = ConsultaHistorico.objects.order_by('-data')[:30]
	if request.method == 'POST':
		if request.POST.get('limpar_historico') == '1':
			ConsultaHistorico.objects.all().delete()
			historico = []
			context = {'resultados': [], 'historico': historico, 'msg': 'Histórico apagado com sucesso!'}
			return render(request, 'consulta/home.html', context)
		logger = logging.getLogger('consulta')
		cnpjs = request.POST.get('cnpjs', '').strip()
		csv_file = request.FILES.get('csv_file')
		tipo = None
		cnpjs_registro = ''

		def on_retry(attempt, wait):
			pass  # Não altera a mensagem, mantém 'Processando...'

		if cnpjs:
			tipo = 'manual'
			cnpj_list, resultados = processar_cnpjs_manualmente(cnpjs, on_retry=on_retry)
			cnpjs_registro = ','.join(cnpj_list)
		elif csv_file:
			# Validação server-side do tipo de upload
			fname_lower = (csv_file.name or '').lower()
			ext = os.path.splitext(fname_lower)[1]
			ctype = (getattr(csv_file, 'content_type', '') or '').lower()
			allowed_ext = getattr(settings, 'ALLOWED_UPLOAD_EXTENSIONS', ['.csv', '.xlsx'])
			allowed_types = getattr(settings, 'ALLOWED_UPLOAD_MIME_TYPES', [])
			if (ext not in allowed_ext) or (allowed_types and ctype not in allowed_types):
				error_msg = 'Arquivo não permitido. Envie apenas CSV ou XLSX.'
			else:
				tipo = 'upload'
				if fname_lower.endswith('.csv'):
					try:
						resultados = processar_csv(csv_file, logger=logger, on_retry=on_retry)
					except Exception as e:
						error_msg = f'Erro ao processar o arquivo: {str(e)}'
				elif fname_lower.endswith('.xlsx'):
					try:
						resultados = processar_xlsx(csv_file, logger=logger, on_retry=on_retry)
					except Exception as e:
						error_msg = f'Erro ao processar o arquivo: {str(e)}'
				else:
					error_msg = 'O arquivo enviado deve ser um CSV ou XLSX.'
		# Limpa status de retry ao fim do processamento
		request.session['status_retry'] = ''
		# Salvar histórico se houver resultados (ou erro); sempre como lista
		if (tipo and (resultados or error_msg)):
			payload_result = resultados if resultados else [{'cnpj': '-', 'nome': '-', 'email': f'Erro: {error_msg}'}]
			ConsultaHistorico.objects.create(
				tipo=tipo,
				cnpjs=cnpjs_registro,
				arquivo_nome=csv_file.name if tipo == 'upload' and csv_file else None,
				resultado=payload_result
			)
		# Salvar resultados atuais na sessão para exportação
		request.session['ultimos_resultados'] = resultados
	context = {'resultados': resultados, 'historico': historico}
	if error_msg:
		context['error_msg'] = error_msg
	return render(request, 'consulta/home.html', context)


@login_required(login_url='login')
def export_resultado_csv(request):
	"""Exporta os últimos resultados da sessão como CSV (sem coluna Data)."""
	resultados = request.session.get('ultimos_resultados', [])
	csv_data = exportar_csv(resultados)
	response = HttpResponse(csv_data, content_type='text/csv')
	response['Content-Disposition'] = 'attachment; filename="resultado.csv"'
	return response


@login_required(login_url='login')
def export_resultado_xlsx(request):
	"""Exporta os últimos resultados da sessão como XLSX (sem coluna Data)."""
	resultados = request.session.get('ultimos_resultados', [])
	xlsx_data = exportar_xlsx(resultados)
	response = HttpResponse(xlsx_data, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
	response['Content-Disposition'] = 'attachment; filename="resultado.xlsx"'
	return response


@login_required(login_url='login')
def export_historico_csv(request):
	"""Exporta todo o histórico do banco como CSV (inclui coluna Data)."""
	historico = ConsultaHistorico.objects.order_by('-data')
	resultados = []
	for h in historico:
		for r in h.resultado:
			r_cpy = r.copy()
			r_cpy['data'] = h.data.strftime('%d/%m/%y')
			resultados.append(r_cpy)
	csv_data = exportar_csv(resultados, include_data=True)
	response = HttpResponse(csv_data, content_type='text/csv')
	response['Content-Disposition'] = 'attachment; filename="historico.csv"'
	return response


@login_required(login_url='login')
def export_historico_xlsx(request):
	"""Exporta todo o histórico do banco como XLSX (inclui coluna Data)."""
	historico = ConsultaHistorico.objects.order_by('-data')
	resultados = []
	for h in historico:
		for r in h.resultado:
			r_cpy = r.copy()
			r_cpy['data'] = h.data.strftime('%d/%m/%y')
			resultados.append(r_cpy)
	xlsx_data = exportar_xlsx(resultados, include_data=True)
	response = HttpResponse(xlsx_data, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
	response['Content-Disposition'] = 'attachment; filename="historico.xlsx"'
	return response


@require_GET
@login_required(login_url='login')
def api_creditos(request):
	"""Retorna créditos CNPJÁ com cache de 24h. Force refresh com ?refresh=1."""
	cache_key = 'cnpja_creditos_v1'
	refresh = request.GET.get('refresh') == '1'
	data = cache.get(cache_key)
	if data is None or refresh:
		try:
			client = CNPJAClient()
			data = client.get_credits(timeout=15)
			# TTL de 24h = 86400s
			cache.set(cache_key, data, timeout=86400)
		except Exception as e:
			if data is None:
				return JsonResponse({'detail': f'Não foi possível obter créditos: {str(e)}'}, status=502)
			# Se já existe cache, retorna cache mesmo em erro
	return JsonResponse(data, safe=False)


def _refresh_creditos_cache_silently():
	"""Atualiza o cache de créditos sem falhar a requisição chamadora."""
	try:
		client = CNPJAClient()
		data = client.get_credits(timeout=15)
		cache.set('cnpja_creditos_v1', data, timeout=86400)
	except Exception:
		pass



@require_GET
@login_required(login_url='login')
def api_detalhes(request, cnpj: str):
    """Retorna o JSON completo salvo para um CNPJ (prioriza sessão atual, depois histórico do banco)."""
    def _digits(s: str) -> str:
        return ''.join(ch for ch in (s or '') if ch.isdigit())

    target = _digits(cnpj)
    if len(target) != 14:
        return JsonResponse({'detail': 'CNPJ inválido'}, status=400)

    # 1) Prioriza resultados do job na sessão (ainda não persistidos)
    job = request.session.get('job') or {}
    for item in (job.get('results') or []):
        if _digits(item.get('cnpj')) == target:
            det = item.get('detalhes')
            if det is not None:
                return JsonResponse(det, safe=False)

    # 2) Procura no histórico do banco (mais recente primeiro)
    qs = ConsultaHistorico.objects.order_by('-data')[:200]
    for h in qs:
        for r in (h.resultado or []):
            try:
                if _digits(r.get('cnpj')) == target and r.get('detalhes') is not None:
                    return JsonResponse(r.get('detalhes'), safe=False)
            except Exception:
                continue

    return JsonResponse({'detail': 'Detalhes não encontrados para este CNPJ.'}, status=404)


class ConsultaCNPJView(APIView):
	"""GET /cnpj/<cnpj>/ retorna o JSON completo da API PRO do CNPJÁ."""
	permission_classes = [IsAuthenticated]
	def get(self, request, cnpj: str):
		s = CNPJQuerySerializer(data={'cnpj': cnpj})
		if not s.is_valid():
			return Response(s.errors, status=status.HTTP_400_BAD_REQUEST)
		cnpj_digits = s.validated_data['cnpj']
		try:
			client = CNPJAClient()
			data = client.get_office(cnpj_digits)
			return Response(data, status=status.HTTP_200_OK)
		except CNPJAClientError as e:
			return Response({ 'detail': str(e) }, status=status.HTTP_400_BAD_REQUEST)
		except Exception:
			return Response({ 'detail': 'Erro interno ao consultar CNPJ' }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


# --------- Abordagem simples com polling (sem Celery) ---------

def _init_job_session(request, items):
	"""Inicializa a estrutura de job na sessão.

	Normaliza `items` (strings ou dicts {cnpj, processo}), aplica deduplicação por par
	(cnpj, processo), e registra contadores/estado para o loop de processamento.
	"""
	# items pode ser lista de strings (cnpj) ou dicts {cnpj, processo}
	normalized = []
	seen = set()  # dedup apenas pares idênticos (cnpj, processo)
	for it in items:
		if isinstance(it, dict):
			c = clean_cnpj(it.get('cnpj', ''))
			p = (it.get('processo') or None)
			key = (c, (p or '')) if c else None
			if c and key not in seen:
				normalized.append({'cnpj': c, 'processo': p})
				seen.add(key)
		else:
			c = clean_cnpj(it)
			key = (c, '') if c else None
			if c and key not in seen:
				normalized.append({'cnpj': c, 'processo': None})
				seen.add(key)
	job = {
		'queue': normalized,
		'processed': 0,
		'total': len(normalized),
		'results': [],
		'status': 'running',  # running | paused | cancelled
		'tipo': None,
		'arquivo_nome': None,
		'cnpjs_str': ','.join([i['cnpj'] for i in normalized]),
	}
	request.session['job'] = job
	request.session.modified = True
	return job


@require_http_methods(["POST"])
@login_required(login_url='login')
def jobs_start(request):
	"""Inicia um job a partir de CSV/XLSX ou lista manual de CNPJs, salvando na sessão."""
	# Use o content-type para decidir como ler o corpo, evitando RawPostDataException
	content_type = (request.META.get('CONTENT_TYPE') or '').lower()
	items = []

	if content_type.startswith('application/json'):
		# Não toque em request.POST/FILES aqui
		try:
			payload = json.loads((request.body or b'{}').decode('utf-8'))
		except Exception:
			payload = {}
		cnpjs_raw = (payload.get('cnpjs') or '').strip()
		if cnpjs_raw:
			items = [c.strip() for c in cnpjs_raw.split(',') if c.strip()]
		job = _init_job_session(request, items)
		job['tipo'] = 'manual'
		request.session['job'] = job
		request.session.modified = True
		return JsonResponse({'total': job['total']})
	else:
		# multipart/form-data ou x-www-form-urlencoded
		cnpjs_raw = (request.POST.get('cnpjs') or '').strip()
		if cnpjs_raw:
			items = [c.strip() for c in cnpjs_raw.split(',') if c.strip()]
		if not items:
			up_file = request.FILES.get('csv_file')
			if up_file:
				fname = (up_file.name or '').lower()
				try:
					if fname.endswith('.xlsx'):
						# Extrai CNPJs do XLSX via openpyxl
						import openpyxl as _openpyxl
						wb = _openpyxl.load_workbook(up_file, read_only=True, data_only=True)
						ws = wb.active
						header_row = next(ws.iter_rows(min_row=1, max_row=1))
						headers = [str(c.value).strip().lower() if c.value else '' for c in header_row]
						keys = ['cnpj', 'cnpj/cpf', 'cnpj_cpf']
						pkeys = ['processo', 'número do processo', 'numero do processo']
						cnpj_idx = None
						proc_idx = None
						for idx, h in enumerate(headers):
							if any(k in h for k in keys):
								cnpj_idx = idx
								break
						for idx, h in enumerate(headers):
							if any(k in h for k in pkeys):
								proc_idx = idx
								break
						if cnpj_idx is not None:
							for row in ws.iter_rows(min_row=2):
								val = row[cnpj_idx].value if cnpj_idx < len(row) else ''
								cnpj_val = clean_cnpj(str(val) if val is not None else '')
								proc_val = (str(row[proc_idx].value).strip() if (proc_idx is not None and row[proc_idx].value) else None)
								if len(cnpj_val) == 14:
									items.append({'cnpj': cnpj_val, 'processo': proc_val})
						else:
							# Fallback: varre todas as células procurando padrões de CNPJ
							pattern = re.compile(r"\d{2}\D?\d{3}\D?\d{3}\D?\d{4}\D?\d{2}")
							for row in ws.iter_rows(min_row=1):
								for cell in row:
									if cell.value is None:
										continue
									text = str(cell.value)
									for m in pattern.findall(text):
										digits = clean_cnpj(m)
										if len(digits) == 14:
											items.append({'cnpj': digits, 'processo': None})
					elif fname.endswith('.csv'):
						# Extrai CNPJs do CSV com fallback de encoding
						raw = up_file.read()
						try:
							decoded = raw.decode('utf-8-sig')
						except UnicodeDecodeError:
							decoded = raw.decode('latin-1')
						import csv as _csv
						import io as _io
						sio = _io.StringIO(decoded)
						# Tenta DictReader primeiro
						try:
							reader = _csv.DictReader(sio)
							keys = ['cnpj', 'CNPJ', 'CNPJ/CPF', 'cnpj_cpf']
							pkeys = ['processo', 'Processo', 'número do processo', 'numero do processo']
							found_by_header = False
							for row in reader:
								matched_this_row = False
								proc_val = None
								for key in row:
									if key is None:
										continue
									if any(k.lower() in key.lower() for k in keys):
										cnpj_val = clean_cnpj(row[key])
										if len(cnpj_val) == 14:
											# captura processo se existir
											for k2 in row:
												if any(pk.lower() in k2.lower() for pk in pkeys):
													proc_val = (row[k2] or '').strip()
													break
											items.append({'cnpj': cnpj_val, 'processo': proc_val})
											matched_this_row = True
											found_by_header = True
								if not matched_this_row:
									# Fallback por linha: vasculha todos os valores
									pattern = re.compile(r"\d{2}\D?\d{3}\D?\d{3}\D?\d{4}\D?\d{2}")
									for v in row.values():
										if not v:
											continue
										for m in pattern.findall(str(v)):
											digits = clean_cnpj(m)
											if len(digits) == 14:
												items.append({'cnpj': digits, 'processo': None})
						except Exception:
							# Se DictReader não funcionar bem (csv caótico), usa csv.reader
							sio.seek(0)
							reader2 = _csv.reader(sio)
							pattern = re.compile(r"\d{2}\D?\d{3}\D?\d{3}\D?\d{4}\D?\d{2}")
							for row in reader2:
								for field in row:
									for m in pattern.findall(str(field)):
										digits = clean_cnpj(m)
										if len(digits) == 14:
											items.append({'cnpj': digits, 'processo': None})
					else:
						return JsonResponse({'detail': 'Tipo de arquivo não suportado. Envie CSV ou XLSX.'}, status=400)
				except Exception as e:
					return JsonResponse({'detail': f'Erro ao ler arquivo: {str(e)}'}, status=400)
		if not items:
			return JsonResponse({'detail': 'Informe cnpjs (JSON/POST) ou envie csv_file.'}, status=400)
		job = _init_job_session(request, items)
		# Define metadados do job conforme origem
		if request.FILES.get('csv_file'):
			job['tipo'] = 'upload'
			job['arquivo_nome'] = request.FILES['csv_file'].name
		else:
			job['tipo'] = 'manual'
		request.session['job'] = job
		request.session.modified = True
		return JsonResponse({'total': job['total']})


@require_http_methods(["POST"])
@login_required(login_url='login')
def jobs_step(request):
	"""Processa um item da fila do job na sessão e retorna o resultado parcial."""
	import time
	from .services import DELAY_SECONDS
	job = request.session.get('job')
	if not job:
		return JsonResponse({'detail': 'Nenhum job em andamento.'}, status=400)
	queue = job.get('queue', [])
	processed = job.get('processed', 0)
	total = job.get('total', 0)
	results = job.get('results', [])
	status_job = job.get('status', 'running')
	if status_job == 'paused':
		return JsonResponse({'status': 'paused', 'processed': processed, 'total': total, 'item': None})
	if status_job == 'cancelled':
		return JsonResponse({'status': 'cancelled', 'processed': processed, 'total': total, 'item': None})
	if processed >= total or not queue:
		return JsonResponse({'status': 'done', 'processed': processed, 'total': total, 'item': None})
	time.sleep(DELAY_SECONDS)
	item = queue.pop(0)
	cnpj = item.get('cnpj') if isinstance(item, dict) else item
	try:
		resultado = consultar_cnpj_api(cnpj)
	except Exception as e:
		resultado = {'cnpj': cnpj, 'nome': '-', 'email': f'Erro: {str(e)}'}
	# anexa processo se existir
	if isinstance(item, dict) and item.get('processo'):
		resultado['processo'] = item['processo']
	results.append(resultado)
	processed += 1
	job.update({'queue': queue, 'processed': processed, 'total': total, 'results': results})
	request.session['job'] = job
	request.session['ultimos_resultados'] = results  # mantém export funcionando
	request.session.modified = True
	return JsonResponse({'status': 'running', 'processed': processed, 'total': total, 'item': resultado})


@require_http_methods(["POST"])
@login_required(login_url='login')
def jobs_finalize(request):
	"""Persiste os resultados do job no histórico e finaliza, limpando o job da sessão."""
	job = request.session.get('job')
	if not job:
		return JsonResponse({'detail': 'Nenhum job em andamento.'}, status=400)
	try:
		tipo = job.get('tipo') or 'manual'
		cnpjs_registro = job.get('cnpjs_str') or ''
		arquivo_nome = job.get('arquivo_nome')
		resultados = job.get('results') or []
		if resultados:
			ConsultaHistorico.objects.create(
				tipo=tipo,
				cnpjs=cnpjs_registro,
				arquivo_nome=arquivo_nome,
				resultado=resultados,
			)
		# limpar job
		request.session.pop('job', None)
		request.session.modified = True
		# tenta atualizar créditos em background (sem bloquear resposta)
		_refresh_creditos_cache_silently()
		return JsonResponse({'status': 'ok'})
	except Exception as e:
		return JsonResponse({'detail': f'Erro ao salvar histórico: {str(e)}'}, status=500)


@require_http_methods(["POST"])
@login_required(login_url='login')
def jobs_pause(request):
	"""Pausa o job em andamento marcando o status como 'paused'."""
	job = request.session.get('job')
	if not job:
		return JsonResponse({'detail': 'Nenhum job em andamento.'}, status=400)
	job['status'] = 'paused'
	request.session['job'] = job
	request.session.modified = True
	return JsonResponse({'status': 'paused'})


@require_http_methods(["POST"])
@login_required(login_url='login')
def jobs_resume(request):
	"""Retoma o job pausado, marcando o status como 'running'."""
	job = request.session.get('job')
	if not job:
		return JsonResponse({'detail': 'Nenhum job em andamento.'}, status=400)
	job['status'] = 'running'
	request.session['job'] = job
	request.session.modified = True
	return JsonResponse({'status': 'running'})


@require_http_methods(["POST"])
@login_required(login_url='login')
def jobs_cancel(request):
	"""Cancela o job atual e esvazia a fila remanescente para encerrar o loop."""
	job = request.session.get('job')
	if not job:
		return JsonResponse({'detail': 'Nenhum job em andamento.'}, status=400)
	job['status'] = 'cancelled'
	# limpa a fila restante para encerrar imediatamente
	job['queue'] = []
	request.session['job'] = job
	request.session.modified = True
	return JsonResponse({'status': 'cancelled'})


# -------------------- Autenticação --------------------

def _client_ip(request):
	"""Obtém o IP do cliente considerando X-Forwarded-For quando presente."""
	xff = request.META.get('HTTP_X_FORWARDED_FOR')
	if xff:
		return xff.split(',')[0].strip()
	return request.META.get('REMOTE_ADDR') or '0.0.0.0'


def login_view(request):
	"""Tela e submissão de login com mitigação de brute force por IP/identificador."""
	if request.user.is_authenticated:
		return redirect('home')
	form = LoginForm(request.POST or None)
	error = None
	ip = _client_ip(request)
	if request.method == 'POST':
		# Axes fará o bloqueio automático; aqui só validamos o formulário
		if form.is_valid():
			ident = form.cleaned_data['username']
			password = form.cleaned_data['password']
			# aceita e-mail ou usuário
			User = get_user_model()
			username = ident
			if '@' in ident:
				try:
					user_obj = User.objects.get(email__iexact=ident)
					username = user_obj.get_username()
				except User.DoesNotExist:
					pass
			user = authenticate(request, username=username, password=password)
			if user is not None:
				auth_login(request, user)
				return redirect('home')
			else:
				messages.error(request, 'Credenciais inválidas.')
		else:
			messages.error(request, 'Preencha os campos corretamente.')
	return render(request, 'auth/login.html', {'form': form})


def logout_view(request):
	"""Efetua logout e redireciona para a tela de login."""
	auth_logout(request)
	return redirect('login')


# signup desabilitado para uso interno (sem rota pública)
